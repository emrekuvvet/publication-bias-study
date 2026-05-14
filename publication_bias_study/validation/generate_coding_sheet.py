"""
generate_coding_sheet.py
------------------------
Draw a stratified random sample of 400 in-scope papers (200 LLM-coded
as in-scope, 200 LLM-coded as out-of-scope from keyword candidates)
and produce human_coding_sheet.xlsx with instructions for RAs.

Usage:
    python validation/generate_coding_sheet.py
    python validation/generate_coding_sheet.py --n 400 --seed 123
"""

import argparse
import pathlib

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

ROOT     = pathlib.Path(__file__).parent.parent.resolve()
IN_PATH  = ROOT / "data" / "classified" / "inscope_papers.parquet"
OUT_PATH = ROOT / "validation" / "human_coding_sheet.xlsx"


INSTRUCTIONS = """\
CODING INSTRUCTIONS FOR RESEARCH ASSISTANTS
============================================

Task: For each paper abstract, code the DIRECTION of the main empirical finding.

Codes:
  +1  The government intervention had a POSITIVE / BENEFICIAL effect
      (e.g., improved market quality, lower cost of capital, higher firm value,
       reduced systemic risk, better investor protection)
  -1  The government intervention had a NEGATIVE / HARMFUL effect
      (e.g., reduced market liquidity, increased costs, unintended consequences,
       welfare loss, market distortion)
   0  NULL, MIXED, or INCONCLUSIVE finding
      (no statistically significant effect, evidence on both sides,
       paper refuses directional stance, purely theoretical)

Rules:
  1. Code the PRIMARY intervention named in the TITLE or first sentence of abstract.
  2. Code the sign of the ESTIMATED COEFFICIENT, not the authors' normative framing.
  3. "Positive regulation is beneficial" is normative framing — ignore it.
  4. Theoretical papers with NO empirical test → code 0.
  5. If you cannot determine direction from abstract alone → code 0 and note.

Ambiguity column:
  Leave blank if confident.
  Write 'Y' if you are uncertain about your code.

Please also code (column 'intervention_type'):
  REG = financial/securities regulation
  BNK = banking/FDIC/bailout policy
  MON = monetary/fiscal policy
  MKT = market microstructure rule
  GOV = corporate governance law
  OTH = other
"""


def draw_sample(df: pd.DataFrame, n: int, seed: int) -> pd.DataFrame:
    in_scope  = df[df["in_scope"] == True]
    out_scope = df[(df["in_scope"] == False) & (df["keyword_hit"] == True)]

    n_in  = min(n // 2, len(in_scope))
    n_out = min(n - n_in, len(out_scope))

    sample = pd.concat([
        in_scope.sample(n=n_in,  random_state=seed),
        out_scope.sample(n=n_out, random_state=seed+1),
    ], ignore_index=True)

    sample = sample.sample(frac=1, random_state=seed+2).reset_index(drop=True)
    return sample


def build_xlsx(sample: pd.DataFrame, out_path: pathlib.Path) -> None:
    wb = Workbook()

    # --- Instructions sheet ---
    ws_inst = wb.active
    ws_inst.title = "Instructions"
    ws_inst.column_dimensions["A"].width = 100
    for i, line in enumerate(INSTRUCTIONS.splitlines(), start=1):
        cell = ws_inst.cell(row=i, column=1, value=line)
        cell.font = Font(name="Courier New", size=10)

    # --- Coding sheet ---
    ws = wb.create_sheet("Coding")

    cols = ["paper_id","source","year","title","abstract",
            "ra1_code","ra1_ambiguity","ra1_intervention_type",
            "ra2_code","ra2_ambiguity","ra2_intervention_type",
            "llm_in_scope","llm_reason"]

    # Widen columns
    widths = {
        "A":15, "B":8, "C":6, "D":50, "E":80,
        "F":10, "G":12, "H":22,
        "I":10, "J":12, "K":22,
        "L":12, "M":50,
    }
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # Header row
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True, size=10)
    for c_idx, col_name in enumerate(cols, start=1):
        cell = ws.cell(row=1, column=c_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    # Data rows
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    alt_fill = PatternFill("solid", fgColor="EBF3FB")

    for r_idx, (_, row) in enumerate(sample.iterrows(), start=2):
        fill = alt_fill if r_idx % 2 == 0 else PatternFill()
        for c_idx, col_name in enumerate(cols, start=1):
            if col_name in row.index:
                val = row[col_name]
                if pd.isna(val):
                    val = ""
            else:
                val = ""
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = border
            cell.fill = fill

    ws.freeze_panes = "A2"
    wb.save(out_path)
    print(f"Saved → {out_path}")
    print(f"  {len(sample):,} papers  ({sample['in_scope'].sum()} in-scope, "
          f"{(~sample['in_scope']).sum()} out-of-scope)")


def main(n: int, seed: int) -> None:
    if not IN_PATH.exists():
        raise FileNotFoundError(
            f"Missing {IN_PATH}. Run 03_classify_inscope.py first."
        )
    df = pd.read_parquet(IN_PATH)
    print(f"Loaded {len(df):,} inscope_papers rows")

    sample = draw_sample(df, n=n, seed=seed)
    build_xlsx(sample, OUT_PATH)


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--n",    type=int, default=400,
                        help="Total sample size (default 400)")
    parser.add_argument("--seed", type=int, default=42)
    args = parser.parse_args()
    main(args.n, args.seed)
